import tkFileDialog as filedialog
import random
import os
from Tkinter import *
from PIL import Image, ImageTk
import glob
#For algorithm
import matlab.engine
import matlab
#For matching and cleanup
from openpyxl import load_workbook
from collections import defaultdict
import csv

eng = matlab.engine.start_matlab()

class App:
    def __init__(self, master):
        master.title('QCA')
        frame = Frame(master)
        frame.grid(row=0,column=0, rowspan=6, sticky=W+E+N+S)
        
        #Will contain the list of images found in the selected directory
        global image_files
        image_files = []
        #If the user has a custom ratio, what is it?
        global custom_ratio
        custom_ratio = StringVar()
        #Does the user have a custom ratio?
        global custom_ratio_bool
        custom_ratio_bool = IntVar()

        def Pull_Extensions():
            #Identifies filetypes entered below
            global filetypes
            filetypes = ex_list.get()
            filetypes = filetypes.split(';')
            print(filetypes)
            T.insert(END,filetypes)
            T.insert(END,'\n')
            
        def Open_Directory():
            #Asks user for directory containing images
            global image_files
            image_files = []
            global image_dir
            image_dir = filedialog.askdirectory(parent=root, initialdir="/", title="Select Directory")
            print(image_dir)
            T.insert(END, image_dir)
            T.insert(END, '\n')

        def Image_Parse():
            #pull individual filenames
            # Grabs all filenames with the correct extensions and puts them in the file list
            for i in range(len(filetypes)):
                file_type_temp = filetypes[i]
                for filename in glob.glob(os.path.join(image_dir, ("*" + str(file_type_temp)))):
                    image_files.append(filename)
            for item in image_files:
                print(item)
                T.insert(END, item)
                T.insert(END, '\n')
        
        def Random_Image():
            #select random image for preview
            global im_select
            im_select = image_files[random.randint(0,(len(image_files)-1))]
            Pho_image = ImageTk.PhotoImage(Image.open(im_select).resize((300,300),Image.ANTIALIAS))
            preview = Label(frame, image = Pho_image)
            preview.image = Pho_image
            preview.grid(row=1, column=1, rowspan = 10)

        def Parse_Parameters():
            print("min cap: " + str(min_cap.get()))
            global min_cap_var
            min_cap_var = int(str(min_cap.get()))
            print("max cap: " + str(max_cap.get()))
            global max_cap_var
            max_cap_var = int(str(max_cap.get()))
            print("min bod: " + str(min_bod.get()))
            global min_bod_var
            min_bod_var = int(str(min_bod.get()))
            print("max bod: " + str(max_bod.get()))
            global max_bod_var
            max_bod_var = int(str(max_bod.get()))
            print("capsule sensitivity: " + str(cap_sens.get()))
            global cap_sens_var
            cap_sens_var = (float(cap_sens.get())/100)
            print("body sensitivity: " + str(bod_sens.get()))
            global bod_sens_var
            bod_sens_var = (float(bod_sens.get())/100)

        def Calculate_Conversion():
            #Getting pixel conversion information from user input
            vtotal = (str(v.get())+str(v2.get()))
            global convert
            if str(custom_ratio_bool.get()) == '1':
                #custom entry
                convert = float(str(custom_ratio.get()))
            elif vtotal == "11":
                #10x with 1x1 binning
                convert = 1.5
            elif vtotal == "12":
                #10x with 2x2 binning
                convert = 0.75
            elif vtotal == "13":
                #10x with 3x3 binning
                convert = 0.5
            elif vtotal == "21":
                #20x with 1x1 binning
                convert = 3.0
            elif vtotal == "22":
                #20x with 2x2 binning
                convert = 1.5
            elif vtotal == "23":
                #20x with 3x3 binning
                convert = 1.0
            elif vtotal == "31":
                #40x with 1x1 binning
                convert = 6.0
            elif vtotal == "32":
                #40x with 2x2 binning
                convert = 3.0
            elif vtotal == "33":
                #40x with 3x3 binning
                convert = 2.0
            elif vtotal == "41":
                #40x (oil) with 1x1 binning
                convert = 6.0
            elif vtotal == "42":
                #40x (oil) with 2x2 binning
                convert = 3.0
            elif vtotal == "43":
                #40x (oil) with 3x3 binning
                convert = 2.0
            elif vtotal == "51":
                #100x with 1x1 binning
                convert = 15.0
            elif vtotal == "52":
                #100x with 2x2 binning
                convert = 7.5
            elif vtotal == "53":
                #100x with 3x3 binning
                convert = 5.0
            else:
                print("nope")
            T.insert(END,(str(convert) + " Pixels = 1um"))
            T.insert(END,'\n')

        def Test_Run():
            #Test out the algorithm on just the random image
            Parse_Parameters()
            test_count = 0
            capsules, bodies = eng.TestRun(im_select, float(min_cap_var), float(max_cap_var), float(cap_sens_var), float(min_bod_var),
                                           float(max_bod_var), float(bod_sens_var), nargout=2)
            for row in bodies:
                bx = row[0]
                by = row[1]
                for row in capsules:
                    cx = row[0]
                    cy = row[1]
                    cr = row[2]
                    if (cx - cr) < bx < (cx + cr) and (cy - cr) < by < (cy + cr):
                        test_count += 1
            T.insert(END,(str(test_count) + " Detected"))
            T.insert(END,'\n')
            Pho_image = ImageTk.PhotoImage(
                Image.open("TestRun.jpeg").resize((300, 300), Image.ANTIALIAS))
            preview = Label(frame, image=Pho_image)
            preview.image = Pho_image
            preview.grid(row=1, column=1, rowspan=10)

        def Analysis():
            Parse_Parameters()
            count = len(image_files)
            filecount = 1
            for item in image_files:
                T.insert(END,"Analyzing image " + str(filecount) + " of " + str(count))
                T.insert(END,'\n')
                eng.Analysis2(str(item), float(min_cap_var), float(max_cap_var), float(cap_sens_var), float(min_bod_var), float(max_bod_var), float(bod_sens_var), str(image_dir), nargout=0)
                filecount += 1
            T.insert(END,"Finished")
            T.insert(END,'\n')

        def Cleanup():
            Parse_Parameters()
            Calculate_Conversion()
            path = (image_dir + '/RawOutput.xlsx')
            wb = load_workbook(path)
            #First remove blank sheets to prevent errors reading data
            if 'Sheet1' in wb.sheetnames:
                std = wb.get_sheet_by_name('Sheet1')
                wb.remove_sheet(std)
            if 'Sheet2' in wb.sheetnames:
                std2 = wb.get_sheet_by_name('Sheet2')
                wb.remove_sheet(std2)
            if 'Sheet3' in wb.sheetnames:
                std3 = wb.get_sheet_by_name('Sheet3')
                wb.remove_sheet(std3)

            global output
            output = defaultdict(list)
            #Go through each sheet to coordinate detected cells
            for sheet in wb:
                global sheet_output
                sheet_output = defaultdict(list)
                bodies = defaultdict(list)
                capsules = defaultdict(list)
                #for each sheet we will then first identify each body and capsule
                body_count = 0
                for row in sheet.iter_rows(row_offset=1):
                    #pulls out each entry as x,y coordinates then radius measurement
                    if row[3].value != None:
                        bodies[body_count].append(float(row[3].value))
                        bodies[body_count].append(float(row[4].value))
                        bodies[body_count].append(float(row[5].value))
                    if row[0].value != None:
                        capsules[body_count].append(float(row[0].value))
                        capsules[body_count].append(float(row[1].value))
                        capsules[body_count].append(float(row[2].value))
                    else:
                        print('no value')
                    body_count = body_count + 1
                # now we need to check to see if the bodies are within a capsule
                for key in bodies:
                    name = (str(sheet) + "-" + str(key))
                    # pull out body centers 1 by 1
                    b_x = float(bodies[key][0])
                    b_y = float(bodies[key][1])
                    b_r = float(bodies[key][2])
                    for key in capsules:
                        # check each body against each capsule
                        c_x = float(capsules[key][0])
                        c_y = float(capsules[key][1])
                        c_r = float(capsules[key][2])
                        # only take the bodies whose center is within a capsule
                        if (c_x - c_r) < b_x < (c_x + c_r) and (c_y - c_r) < b_y < (c_y + c_r):
                            # Here is where we store the relevant data
                            output[name].append(name)
                            output[name].append((c_r/convert))
                            output[name].append(c_x)
                            output[name].append(c_y)
                            output[name].append((b_r / convert))
                            output[name].append(((c_r/convert)-(b_r / convert)))
                del sheet_output
            #Detected bodies withing capsules are output to a csv file with all relevant info
            pathout = (image_dir + '/CleanedOutput.csv')
            with open(pathout, 'wt') as outfile:
                fields = ['Image File', 'Total Radius (um)', 'Capsule x', 'Capsule y', 'Body Radius (um)', 'Capsule Radius (um)']
                writer = csv.writer(outfile, delimiter=',')
                writer.writerow(fields)
                for key in output:
                    writer.writerow(output[key])

        def Pull_Bodies():
            Parse_Parameters()
            Calculate_Conversion()
            path = (image_dir + '/RawOutput.xlsx')
            wb = load_workbook(path)
            # First remove blank sheets to prevent errors reading data
            if 'Sheet1' in wb.sheetnames:
                std = wb.get_sheet_by_name('Sheet1')
                wb.remove_sheet(std)
            if 'Sheet2' in wb.sheetnames:
                std2 = wb.get_sheet_by_name('Sheet2')
                wb.remove_sheet(std2)
            if 'Sheet3' in wb.sheetnames:
                std3 = wb.get_sheet_by_name('Sheet3')
                wb.remove_sheet(std3)
            global output
            output = defaultdict(list)
            # Go through each sheet to coordinate detected cells
            for sheet in wb:
                global sheet_output
                sheet_output = defaultdict(list)
                bodies = defaultdict(list)
                # for each sheet we will then first identify each body and capsule
                body_count = 0
                for row in sheet.iter_rows(row_offset=1):
                    # pulls out all the bodies if there is data in the row
                    if row[3].value != None:
                        bodies[body_count].append(float(row[5].value))
                    else:
                        print('no value')
                    body_count = body_count + 1
                # now we need to check to see if the bodies are within a capsule
                for key in bodies:
                    # keeping this as a separate for because the naming scheme is set up
                    name = (str(sheet) + "-" + str(key))
                    # pull out body centers 1 by 1
                    b_r = float(bodies[key][0])
                    output[name].append(name)
                    output[name].append((b_r / convert))
                del sheet_output
            # Detected bodies withing capsules are output to a csv file with all relevant info
            pathout = (image_dir + '/CleanedBodies.csv')
            with open(pathout, 'wt') as outfile:
                fields = ['Image File', 'Body Radius (um)']
                writer = csv.writer(outfile, delimiter=',')
                writer.writerow(fields)
                for key in output:
                    writer.writerow(output[key])

        def Pull_Capsules():
            Parse_Parameters()
            Calculate_Conversion()
            path = (image_dir + '/RawOutput.xlsx')
            wb = load_workbook(path)
            # First remove blank sheets to prevent errors reading data
            if 'Sheet1' in wb.sheetnames:
                std = wb.get_sheet_by_name('Sheet1')
                wb.remove_sheet(std)
            if 'Sheet2' in wb.sheetnames:
                std2 = wb.get_sheet_by_name('Sheet2')
                wb.remove_sheet(std2)
            if 'Sheet3' in wb.sheetnames:
                std3 = wb.get_sheet_by_name('Sheet3')
                wb.remove_sheet(std3)
            global output
            output = defaultdict(list)
            # Go through each sheet to coordinate detected cells
            for sheet in wb:
                global sheet_output
                sheet_output = defaultdict(list)
                capsules = defaultdict(list)
                # for each sheet we will then first identify each body and capsule
                body_count = 0
                for row in sheet.iter_rows(row_offset=1):
                    # pulls out all the bodies if there is data in the row
                    if row[0].value != None:
                        capsules[body_count].append(float(row[2].value))
                    else:
                        print('no value')
                    body_count = body_count + 1
                # now we need to check to see if the bodies are within a capsule
                for key in capsules:
                    # keeping this as a separate for because the naming scheme is set up
                    name = (str(sheet) + "-" + str(key))
                    # pull out body centers 1 by 1
                    c_r = float(capsules[key][0])
                    output[name].append(name)
                    output[name].append((c_r / convert))
                del sheet_output
            # Detected bodies withing capsules are output to a csv file with all relevant info
            pathout = (image_dir + '/CleanedCapsules.csv')
            with open(pathout, 'wt') as outfile:
                fields = ['Image File', 'Total Radius (um)']
                writer = csv.writer(outfile, delimiter=',')
                writer.writerow(fields)
                for key in output:
                    writer.writerow(output[key])

        #Make row locations dynamic without having to update each time we fuck around with code
        global row_count
        row_count = 0

        Label(frame, text="Step 1 - Input Image Filetype Extensions", fg="red", justify=LEFT, padx=20).grid(row=row_count,column=0)
        row_count += 1

        #Asking for file extensions of images
        Label(frame, text="Image File Extension(s) in the format of: .TIF" + "\n" + "(separate by semicolon)", justify=LEFT, padx=20).grid(row=row_count, column=0)
        row_count += 1
        global ex_list
        ex_list = StringVar()
        e1 = Entry(frame, textvariable=ex_list)
        e1.grid(row=row_count, column=0)
        row_count += 1
        ex_grab = Button(frame, text="Enter", command=Pull_Extensions)
        ex_grab.grid(row=row_count, column=0)
        row_count += 1

        Label(frame, text="Step 2 - Open Directory Containing Image Files", fg="red", padx=20).grid(row=row_count,column=0)
        row_count += 1

        #Opening the directory
        open_button = Button(frame, text="Select Directory", command=Open_Directory)
        open_button.grid(row=row_count, column=0)
        row_count += 1

        Label(frame, text="Step 3 - Generate Image List", fg="red", justify=LEFT, padx=20).grid(row=row_count,column=0)
        row_count += 1
        image_lister = Button(frame, text="Generate Image List", command=Image_Parse)
        image_lister.grid(row=row_count,column=0)
        row_count += 1
        Label(frame, text="Ensure list to the right is accurate", justify=LEFT, padx=20).grid(row=row_count,column=0)
        row_count += 1

        #Generate list of images to make sure everything worked okay
        T = Text(frame)
        T.grid(row=1,column=2, rowspan=10, columnspan=2)

        Label(frame, text="Step 4 - Preview Random Image from List", fg="red", justify=LEFT, padx=20).grid(row=row_count,column=0)
        row_count += 1
        rand_gen = Button(frame, text="Select Random Image", command=Random_Image)
        rand_gen.grid(row=row_count,column=0)
        row_count += 1

        Label(frame, text="Step 5 - Input Microscopy Parameters", fg="red", justify=LEFT, padx=20).grid(row=row_count,column=0)
        row_count += 1
        #Selecting microscope objective information for images
        global v
        v = IntVar()
        v.set(1)
        Objectives = [("10x",1),("20x",2),("40x",3),("40x Oil",4),("100x Oil",5)]
        Label(frame, text="Choose your objective: ", justify=LEFT, padx=20).grid(row=row_count,column=0)
        temp_row_count = row_count
        row_count += 1
        for txt, val in Objectives:
            Radiobutton(frame, text=txt, padx=20, variable=v, value=val).grid(row=row_count,column=0)
            row_count += 1
        #Add an option for custom conversion
        custom_convert = Checkbutton(frame, text="Custom Pixel Conversion", variable=custom_ratio_bool)
        custom_convert.grid(row=temp_row_count, column=2)
        #Entry field for custom button
        custom_convert_entry = Entry(frame, textvariable=custom_ratio)
        custom_convert_entry.grid(row=(temp_row_count+1), column=2)
        #Add a button for pixel um conversion
        conversion = Button(frame, text="Calculate Conversion", command=Calculate_Conversion)
        conversion.grid(row=(temp_row_count+2), column=2)
        #Selecting Binning information for images
        global v2
        v2 = IntVar()
        v2.set(1)
        Binning = [("1x1",1),("2x2",2),("3x3",3)]
        Label(frame, text="Binning?", justify=LEFT, padx=20).grid(row=temp_row_count,column=1)
        temp_row_count += 1
        for txt, val in Binning:
            Radiobutton(frame, text=txt, padx=20, variable=v2, value=val).grid(row=temp_row_count,column=1)
            temp_row_count += 1

        Label(frame, text="Step 6a - Input Detection Algorithm Parameters", fg="red", justify=LEFT, padx=20).grid(row=row_count,column=0)
        row_count += 1
        
        #Asking for circle detection parameters
        Label(frame, text="Min Capsule Radius (pixels)", justify=LEFT, padx=20).grid(row=row_count, column=0)
        Label(frame, text="Max Capsule Radius (pixels)", justify=LEFT, padx=20).grid(row=row_count, column=1)
        Label(frame, text="Min Cell Body Radius (pixels)", justify=LEFT, padx=20).grid(row=row_count, column=2)
        Label(frame, text="Max Cell Body Radius (pixels)", justify=LEFT, padx=20).grid(row=row_count, column=3)
        row_count += 1
        global min_cap
        min_cap = StringVar()
        e2 = Entry(frame, textvariable=min_cap)
        e2.insert(END,str(10))
        e2.grid(row=row_count, column=0)
        global max_cap
        max_cap = StringVar()
        e3 = Entry(frame, textvariable=max_cap)
        e3.insert(END,str(45))
        e3.grid(row=row_count, column=1)
        global min_bod
        min_bod = StringVar()
        e4 = Entry(frame, textvariable=min_bod)
        e4.insert(END,str(4))
        e4.grid(row=row_count, column=2)
        global max_bod
        max_bod = StringVar()
        e5 = Entry(frame, textvariable=max_bod)
        e5.insert(END,str(30))
        e5.grid(row=row_count, column=3)
        row_count += 1
        
        Label(frame, text="Capsule Sensitivity", justify=LEFT, padx=20).grid(row=row_count, column=0)
        Label(frame, text="Cell Body Sensitivity", justify=LEFT, padx=20).grid(row=row_count, column=1)
        row_count += 1
        global cap_sens
        cap_sens = IntVar()
        s1 = Scale(frame, variable=cap_sens, from_=0, to=100, orient=HORIZONTAL)
        s1.set(82)
        s1.grid(row=row_count,column=0)
        global bod_sens
        bod_sens = IntVar()
        s2 = Scale(frame, variable=bod_sens, from_=0, to=100, orient=HORIZONTAL)
        s2.set(80)
        s2.grid(row=row_count,column=1)
        row_count += 1

        #Random Image Test Run
        Label(frame, text="Step 6b - Test Parameters", fg="red", justify=LEFT, padx=20).grid(row=row_count,column=0)
        row_count += 1
        test_run = Button(frame, text="Run Test", command=Test_Run)
        test_run.grid(row=row_count,column=0)
        row_count += 1
        
        #Running the Algorithm for real
        Label(frame, text="Step 7a - Run the Algorithm on the Full Image List", fg="red", justify=LEFT, padx=20).grid(row=row_count,column=0)
        temp_row_count2 = row_count
        row_count += 1
        alg_run = Button(frame, text="Begin Analysis", command=Analysis)
        alg_run.grid(row=row_count,column=0)
        row_count += 1

        #Matching Bodies to Capsules
        Label(frame, text="Step 7b - Match Cell Bodies to Capsules, Cleanup, and Calculate", fg="red", justify=LEFT, padx=20).grid(row=temp_row_count2,column=1)
        Label(frame, text="Alternate Uses", fg="red", justify=LEFT, padx=20).grid(row=temp_row_count2, column=2)
        temp_row_count2 += 1
        matching_run = Button(frame, text="Match and Cleanup", command=Cleanup)
        matching_run.grid(row=temp_row_count2,column=1)
        body_run = Button(frame, text="Only Pull Bodies", command=Pull_Bodies)
        body_run.grid(row=temp_row_count2, column=2)
        temp_row_count2 += 1
        capsule_run = Button(frame, text="Only Pull Capsules", command=Pull_Capsules)
        capsule_run.grid(row=temp_row_count2, column=2)

        Label(frame, text='Data is saved to the image folder as \'CleanedOutput.csv\'').grid(row=(temp_row_count2-1), column=3)
        

root = Tk()
app = App(root)
root.mainloop()